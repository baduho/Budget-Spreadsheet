from openpyxl import Workbook
from openpyxl import load_workbook
from calendar import monthrange
from openpyxl.chart import LineChart
from openpyxl.chart.axis import DateAxis

MonthIndex = {"JAN":1,"FEB":2,"MAR":3,"APR":4,"MAY":5,"JUN":6,"JUL":7,"AUG":8,"SEP":9,"OCT":10,"NOV":11,"DEC":12}

Restart = "Y"
while(Restart == "Y"):
    CreateNew = input("Do you want to create a new Sheet (N) or add Transactions to an old one (O)?")
    '''The program has 2 main functions: creating a new worksheet and adding transactions to it or opening an existing sheet and adding transactions to it'''

    '''This block of code makes sure that the the input given is one of the 2 options (N or O)
    error is 1 when the input is invalid and will be changed to 0 when a valid input is detected'''
    error = 1
    while(error == 1):
        if(CreateNew in ["N", "O"]):
            error = 0
        else:
            print("Invalid Input! Try Again")
            CreateNew = input("Do you want to create a new Sheet (N) or add Transactions to an old one (O)?")
        

    #CHECKS IF THE WORKBOOK ALREADY EXISTS
    import os
    exists = os.path.exists('Monthly_Transactions.xlsx')

    #CREATING A NEW SHEET/WORKBOOK
    if(CreateNew == "N"):
        
        #IF THE WORKBOOK DOES NOT EXIST IT WILL CREATE A NEW ONE, IF IT EXISTS A NEW SHEET WILL BE ADDED 
        if(exists == True):
            wb = load_workbook(filename = 'Monthly_Transactions.xlsx')
            print(wb.sheetnames)
            OpenSheet = input("Which Month and Year?(e.g. JAN20)")

            '''The input for OpenSheet has to have the first three characters in capital letters and representing a month followed by 2 numbers representing the year
            error is at first set to 1 to indicate a invalid input. If a valid input is detected error is set to 0 to exit the loop'''
            error = 1
            while(error == 1): 
                try:
                    int(OpenSheet[3:])
                    error = 0
                except:
                    print("Invalid Input! The first three letters should represent a month and the following number a year")
                    OpenSheet = input("Which Month and Year?(e.g. JAN20)")
                
            error = 1
            while(error == 1):
                if(((OpenSheet[0:3] in MonthIndex) == True) and (len(OpenSheet) == 5)):
                    error = 0
                else:
                    print("Invalid Input! The first three letters should represent a month and the following number a year")
                    OpenSheet = input("Which Month and Year?(e.g. JAN20)")
            

            '''When a new sheet will be added and at least one sheet exists already the program has to get the ending balance from the previous sheet
            to set the starting balance on the new sheet equal to the previous ending balance'''
            ws = wb[wb.sheetnames[len(wb.sheetnames)-1]]
            StartingBalance = ws["C"+str(ws.max_row)].value #problem how to make the formula work when rows added afterwards
            wb.create_sheet(title=OpenSheet)
            ws = wb[OpenSheet]   
            
        if(exists == False):
            wb = Workbook()
            ws =  wb.active
            OpenSheet = input("Which Month and Year?(e.g. JAN20)")
            
            '''The input for OpenSheet has to have the first three characters in capital letters and representing a month followed by 2 numbers representing the year
            error is at first set to 1 to indicate a invalid input. If a valid input is detected error is set to 0 to exit the loop'''
            error = 1
            while(error == 1): 
                try:
                    int(OpenSheet[3:])
                    error = 0
                except:
                    print("Invalid Input! The first three letter should represent a month and the following number a year")
                    OpenSheet = input("Which Month and Year?(e.g. JAN20)")
                
            error = 1
            while(error == 1):
                if(((OpenSheet[0:3] in MonthIndex) == True) and (len(OpenSheet) == 5)):
                    error = 0
                else:
                    print("Invalid Input! The first three letter should represent a month and the following number a year")
                    OpenSheet = input("Which Month and Year?(e.g. JAN20)")
            
            ws.title = OpenSheet
            StartingBalance = input("Starting Balance?")

            '''Starting balance has to be a number. This part checks if the starting balance can be converted into a float
            and if not it tells you your starting balance is invalid and asks you to input the starting balance again'''
            error = 1
            while(error == 1):
                try:
                    float(StartingBalance)
                    error = 0
                except:
                    print("Invalid Input! Try Again")
                    StartingBalance = input("Starting Balance?")
            StartingBalance = float(StartingBalance)
                

        '''Creating the headers in the first row of every column we need'''  
        ws["A1"] = "Date"
        ws["B1"] = "Item Description"
        ws["C1"] = "Amount"
        ws["D1"] = "Chg%"
        '''Creating the first transaction that is the starting balance according to user input or previous ending balance depeding if it is the first sheet or not.
        The Date is always set to the first day of the month'''
        ws["B2"] = "Starting Balance"
        ws["C2"] = StartingBalance
        ws["A2"] = "=DATE("+str(OpenSheet[3:5])+","+str(MonthIndex[OpenSheet[0:3]])+",01)"
        '''The ending balance is created by summing up all the transaction of the month including the starting balance the date is always set to the last day of the month.
        Ending balance is always the last row of input in the sheet'''
        ws["A"+str(ws.max_row + 1)] = "=EOMONTH(A2,0)"
        ws["B"+str(ws.max_row)] = "Ending Balance"
        ws["C"+str(ws.max_row)] ="=SUM('"+OpenSheet+"'!C2:C"+str(ws.max_row - 1)+")"
        '''Setting the width of the columns based on the anticipated lengh of the inputs'''
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 20


    #IF YOU WANT TO ADD TRANSACTIONS TO AN EXISTING WORKBOOK BUT THE WORKBOOK DOES NOT EXIST IT ENDS THE PROGRAM
    if(CreateNew == "O"):
        if(exists==False):
            print("You don't have a Workbook yet!")
            exit()
        else:
            wb = load_workbook(filename = 'Monthly_Transactions.xlsx')
            print(wb.sheetnames)
            OpenSheet = input("Which Sheet do you want to open?")

            '''Checking if the sheet you are trying to access exists in the workbook.
            If the name of the sheet you are entering as an input exist as a worksheet name the code sets the error variable to 0 to exit the loop'''
            error = 1
            while(error == 1):
                if(OpenSheet in wb.sheetnames):
                    error = 0
                else:
                    print("Invalid Input! Try Again")
                    OpenSheet = input("Which Sheet do you want to open?")

            ws = wb[OpenSheet]

    #ADDING TRANSACTIONS
    if(CreateNew == "O"):
        NewTransaction = "Y"
    if(CreateNew == "N"):
        NewTransaction = input("Do you want to add a Transaction? (Y/N)")

        '''The input for adding a new transaction has to be Y or N. Error variable will be changed to 0 and the loop will be exited only if the given input is a capital Y or N'''
        error = 1
        while(error == 1):
            if(NewTransaction in ["Y", "N"]):
                error = 0
            else:
                print("Invalid Input! Try Again")
                NewTransaction = input("Do you want to add a Transaction? (Y/N)")
                

    '''This part is extracting the month and the year for the dates from the inputted sheetname'''
    Month = MonthIndex[OpenSheet[0:3]]
    Year = OpenSheet[3:5]
    MonthRange = monthrange(2000+int(Year),Month)
    DateRange = list(range(1,MonthRange[1]+1))
   
    while(NewTransaction == "Y"):
        Date = input("Date of the Transaction (Day)")

        '''Checking that the inputted date is really a date in the chosen month. For example if the month is February you can't input a transaction on the 31st'''
        error = 1
        while(error == 1):
            while(error == 1):
                try:
                    int(Date)
                    error = 0
                except:
                    print("Invalid Input! Try Again")
                    Date = input("Date of the Transaction (Day)")
            error = 1    
            if((int(Date) in DateRange) and (len(Date)<=2)):
                error = 0
            else:
                print("Invalid Input! Try Again")
                Date = input("Date of the Transaction (Day)")
        
        Item = input("Insert Item Description")
        Price = input("Price")

        '''Cheking that the inputted price is really a number by trying to convert it into a float'''
        error = 1
        while(error == 1):
            try:
                float(Price)
                error = 0
            except:
                print("Invalid Input! Try Again")
                Price = input("Price")
        Price = float(Price)       
            
        ws["A"+str(ws.max_row)] = "=DATE("+str(Year)+","+str(Month)+","+Date[0:2]+")"
        ws["B"+str(ws.max_row)] = Item
        ws["C"+str(ws.max_row)] = Price
        ws["A"+str(ws.max_row + 1)] = "=EOMONTH(A2,0)"
        ws["B"+str(ws.max_row)] = "Ending Balance"
        ws["C"+str(ws.max_row)] ="=SUM('"+OpenSheet+"'!C2:C"+str(ws.max_row - 1)+")"

        #PERCENTAGES
        starting_balance = ws["C2"].value
        ending_balance = ws["C"+str(ws.max_row)].value


        a = ending_balance[1:]

        if(OpenSheet == wb.sheetnames[0]): 
            b = starting_balance
        else:
            b = starting_balance[1:]
                
        ws["D"+str(ws.max_row)] = "=IF("+str(a)+"/"+str(b)+"<0,\"N/A\","+str(a)+"/"+str(b)+"-1)"
        ws["D"+str(ws.max_row-1)] = ""
        
        
        NewTransaction = input("Do you want to add another Transaction? (Y/N)")

    #FORMATTING CELLS    
    for row in range(2,ws.max_row + 1):
        ws.cell(row=row,column=1).number_format = "dd/mm/yy"
    for row in range(2,ws.max_row + 1):
        ws.cell(row=row,column=3).number_format = "0.00"
    ws["D"+str(ws.max_row)].number_format = "0.00%"
    ws.auto_filter.ref = "A1:C"+str(ws.max_row)

    #CREATING THE COLUMNS WITH THE BALANCE AFTER EVERY TRANSACTION
    ws["F2"] = ws["C2"].value
    for i in range(3,ws.max_row):
        ws["F"+str(i)] = "=F"+str(i-1)+"+C"+str(i)
    ws["F"+str(ws.max_row)] = ws["C"+str(ws.max_row)].value
    
    #PLOTTING THE CHART
    c1 = LineChart()
    c1.title = "Balance"
    c1.style = 2 #CAN BE CHANGED TO WHAT EVER BETWEEN 1 AND 48 TO CHANGE THE STYLE OF THE START
    c1.y_axis.title = "$$$$"
    c1.x_axis.number_format = 'dd'
    c1.x_axis.majorTimeUnit = "days"
    c1.x_axis.title = "Day"
    c1.legend = None
    c1.add_data(OpenSheet+"!F2:F"+str(ws.max_row))
    c1.set_categories(OpenSheet+"!A2:A"+str(ws.max_row))

    ws.add_chart(c1, "F1")

        
    #SAVES THE WORKBOOK
    wb.save("Monthly_Transactions.xlsx")

    Restart = input("Do you still want to add a new Sheet or add Transactions to an existing one?(Y/N)")

    error = 1
    while(error == 1):
        if(Restart in ["Y","N"]):
            error = 0
        else:
            print("Invalid Input! Try Again")
            Restart = input("Do you still want to add a new Sheet or add Transactions to an existing one?(Y/N)")
            
